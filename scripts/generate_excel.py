import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def create_styled_excel(filepath, title, sheets_data):
    """
    filepath: path to save .xlsx
    title: Report Title
    sheets_data: dict of sheet_name -> list of rows (dict or list)
                 e.g. {"Backend Tests": [{"id": "TC-01", "name": "...", "status": "PASSED", ...}]}
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default sheet

    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    pass_font = Font(name="Segoe UI", size=10, color="006100", bold=True)
    
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fail_font = Font(name="Segoe UI", size=10, color="9C0006", bold=True)
    
    warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    warn_font = Font(name="Segoe UI", size=10, color="9C6500", bold=True)

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    for sheet_name, rows in sheets_data.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        
        if not rows:
            ws.append(["No Data Recorded"])
            continue

        # Check if rows are dicts or lists
        if isinstance(rows[0], dict):
            headers = list(rows[0].keys())
            headers_formatted = [h.replace("_", " ").title() for h in headers]
            ws.append(headers_formatted)
            
            for row_idx, data_dict in enumerate(rows, start=2):
                row_vals = [data_dict.get(k, "") for k in headers]
                ws.append(row_vals)
                
                # Format status cell if 'status' column present
                for col_idx, h in enumerate(headers, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    val_str = str(cell.value).upper()
                    if h.lower() in ("status", "result", "severity"):
                        if val_str in ("PASSED", "PASS", "SUCCESS", "LOW", "INFO"):
                            cell.fill = pass_fill
                            cell.font = pass_font
                        elif val_str in ("FAILED", "FAIL", "CRITICAL", "HIGH"):
                            cell.fill = fail_fill
                            cell.font = fail_font
                        elif val_str in ("WARNING", "MEDIUM", "SKIPPED"):
                            cell.fill = warn_fill
                            cell.font = warn_font
        else:
            headers = rows[0]
            ws.append(headers)
            for row in rows[1:]:
                ws.append(row)

        # Style header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Column width adjustment
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 65)

    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    print(f"[EXCEL] Generated report successfully at: {filepath}")

if __name__ == "__main__":
    sample_data = {
        "Sample Sheet": [
            {"test_id": "TC-001", "name": "Sample Test", "status": "PASSED", "details": "Execution clean"}
        ]
    }
    create_styled_excel("Test_Results/Sample_Report.xlsx", "Sample Report", sample_data)
