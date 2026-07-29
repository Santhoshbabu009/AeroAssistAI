import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.chart import PieChart, Reference
from datetime import datetime

class SeleniumExcelReporter:
    def __init__(self, output_path=None):
        if output_path is None:
            base_dir = os.path.dirname(os.path.dirname(__file__))
            date_str = datetime.now().strftime('%Y%m%d_%H%M%S')
            os.makedirs(os.path.join(base_dir, 'reports', 'excel'), exist_ok=True)
            self.output_path = os.path.join(base_dir, 'reports', 'excel', f'selenium_report_{date_str}.xlsx')
        else:
            self.output_path = output_path
            
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Test Results"
        
        headers = [
            "Test ID", "Module", "Feature", "Requirement ID", "Priority", 
            "Severity", "Environment", "Browser", "Device", "Platform", 
            "Precondition", "Test Steps", "Expected Result", "Actual Result", 
            "Status", "Execution Time (s)", "Retry Count", "Executed By", 
            "Execution Date", "Screenshot Path", "Evidence", "Remarks"
        ]
        
        self.ws.append(headers)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        
        for cell in self.ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        self.results = {"PASS": 0, "FAIL": 0, "SKIPPED": 0, "BLOCKED": 0}

    def add_result(self, result_dict):
        row = [
            result_dict.get("Test ID", ""),
            result_dict.get("Module", ""),
            result_dict.get("Feature", ""),
            result_dict.get("Requirement ID", ""),
            result_dict.get("Priority", "Medium"),
            result_dict.get("Severity", "Medium"),
            result_dict.get("Environment", "QA"),
            result_dict.get("Browser", "Chrome"),
            result_dict.get("Device", "Desktop"),
            result_dict.get("Platform", "Windows"),
            result_dict.get("Precondition", ""),
            result_dict.get("Test Steps", ""),
            result_dict.get("Expected Result", ""),
            result_dict.get("Actual Result", ""),
            result_dict.get("Status", "PASS"),
            result_dict.get("Execution Time (s)", 0),
            result_dict.get("Retry Count", 0),
            result_dict.get("Executed By", "Automation"),
            result_dict.get("Execution Date", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            result_dict.get("Screenshot Path", ""),
            result_dict.get("Evidence", ""),
            result_dict.get("Remarks", "")
        ]
        
        self.ws.append(row)
        current_row = self.ws.max_row
        status = result_dict.get("Status", "PASS").upper()
        
        if status in self.results:
            self.results[status] += 1
            
        colors = {
            "PASS": "22C55E",
            "FAIL": "EF4444",
            "SKIPPED": "F97316",
            "BLOCKED": "A855F7"
        }
        
        fill_color = colors.get(status, "FFFFFF")
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        for cell in self.ws[current_row]:
            if cell.column == 15: # Status column
                cell.fill = fill
                cell.font = Font(bold=True)

    def save(self):
        # Auto-size columns
        for col in self.ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            self.ws.column_dimensions[column].width = min(adjusted_width, 50)
            
        self.ws.freeze_panes = "A2"
        self.ws.auto_filter.ref = self.ws.dimensions
        
        # Summary Sheet
        summary_ws = self.wb.create_sheet("Summary", 0)
        summary_ws.append(["Metric", "Count"])
        
        total = sum(self.results.values())
        summary_ws.append(["Total Tests", total])
        summary_ws.append(["Passed", self.results["PASS"]])
        summary_ws.append(["Failed", self.results["FAIL"]])
        summary_ws.append(["Skipped", self.results["SKIPPED"]])
        summary_ws.append(["Blocked", self.results["BLOCKED"]])
        
        pass_pct = (self.results["PASS"] / total * 100) if total > 0 else 0
        summary_ws.append(["Pass %", f"{pass_pct:.2f}%"])
        
        chart = PieChart()
        labels = Reference(summary_ws, min_col=1, min_row=3, max_row=6)
        data = Reference(summary_ws, min_col=2, min_row=2, max_row=6)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.title = "Execution Results"
        
        summary_ws.add_chart(chart, "D2")
        
        os.makedirs(os.path.dirname(self.output_path), exist_ok=True)
        self.wb.save(self.output_path)
