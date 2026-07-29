import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from typing import List, Dict, Any

class SecurityExcelReporter:
    def __init__(self, filepath: str = "security_report.xlsx"):
        self.filepath = filepath
        self.results = []
    
    def add_result(self, test_id: str, category: str, owasp_ref: str, severity: str, endpoint: str, payload: str, expected: str, actual: str, status: str, cve: str, remediation: str, evidence: str):
        self.results.append({
            "Test ID": test_id,
            "Vulnerability Category": category,
            "OWASP Reference": owasp_ref,
            "Severity": severity,
            "Endpoint Tested": endpoint,
            "Payload Used": payload,
            "Expected Result": expected,
            "Actual Result": actual,
            "Status": status,
            "CVE Reference": cve,
            "Remediation": remediation,
            "Evidence": evidence
        })
    
    def generate_report(self):
        df = pd.DataFrame(self.results)
        with pd.ExcelWriter(self.filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Security Results', index=False)
            workbook = writer.book
            worksheet = writer.sheets['Security Results']
            
            # Apply color coding
            color_map = {
                "CRITICAL": "FF0000",
                "HIGH": "FFA500",
                "MEDIUM": "FFFF00",
                "LOW": "0000FF",
                "PASS": "00FF00",
                "FAIL": "FF0000"
            }
            
            for row in range(2, len(df) + 2):
                severity = worksheet[f"D{row}"].value
                status = worksheet[f"I{row}"].value
                
                if severity in color_map:
                    worksheet[f"D{row}"].fill = PatternFill(start_color=color_map[severity], end_color=color_map[severity], fill_type="solid")
                if status in color_map:
                    worksheet[f"I{row}"].fill = PatternFill(start_color=color_map[status], end_color=color_map[status], fill_type="solid")
            
            # Create Summary Sheet
            summary_df = df['Severity'].value_counts().reset_index()
            summary_df.columns = ['Severity', 'Count']
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
