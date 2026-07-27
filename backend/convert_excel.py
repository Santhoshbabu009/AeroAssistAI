import openpyxl
import html

wb = openpyxl.load_workbook(r'c:\Users\santh\AndroidStudioProjects\AeroAssistAI\backend\AeroAssist_Backend_Security_Report.xlsx')

html_content = """<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>AeroAssist Security Audit Report</title>
    <style>
        body { font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; margin: 20px; background: #0f172a; color: #f8fafc; }
        h1 { color: #38bdf8; text-align: center; font-size: 24px; margin-bottom: 5px; }
        p.subtitle { text-align: center; color: #94a3b8; font-size: 14px; margin-bottom: 25px; }
        .tab-bar { display: flex; gap: 10px; justify-content: center; margin-bottom: 20px; flex-wrap: wrap; }
        .tab-btn { background: #1e293b; color: #94a3b8; border: 1px solid #334155; padding: 10px 20px; border-radius: 8px; cursor: pointer; font-weight: 600; }
        .tab-btn.active { background: #0284c7; color: white; border-color: #38bdf8; }
        .sheet-container { display: none; }
        .sheet-container.active { display: block; }
        table { width: 100%; border-collapse: collapse; background: #1e293b; border-radius: 8px; overflow: hidden; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.5); }
        th { background: #0f172a; color: #38bdf8; text-align: left; padding: 12px; font-size: 13px; border-bottom: 2px solid #334155; }
        td { padding: 10px 12px; border-bottom: 1px solid #334155; font-size: 13px; color: #e2e8f0; vertical-align: top; }
        tr:hover { background: #334155; }
        .badge-pass { background: #15803d; color: #4ade80; padding: 4px 8px; border-radius: 4px; font-weight: bold; font-size: 11px; }
    </style>
</head>
<body>
    <h1>🛡️ AeroAssist AI - Backend Security & Test Report</h1>
    <p class="subtitle">Generated Comprehensive Security Test Matrix (463 Test Cases)</p>
    <div class="tab-bar">
"""

sheet_ids = []
for idx, name in enumerate(wb.sheetnames):
    s_id = f"sheet_{idx}"
    sheet_ids.append((s_id, name))
    active_cls = " active" if idx == 0 else ""
    html_content += f'<button class="tab-btn{active_cls}" onclick="showSheet(\'{s_id}\', this)">{html.escape(name)}</button>\n'

html_content += "</div>\n"

for idx, (s_id, name) in enumerate(sheet_ids):
    active_cls = " active" if idx == 0 else ""
    ws = wb[name]
    html_content += f'<div id="{s_id}" class="sheet-container{active_cls}"><table>\n'
    first = True
    for row in ws.iter_rows(values_only=True):
        if not any(row): continue
        html_content += "<tr>\n"
        tag = "th" if first else "td"
        for cell in row:
            val = str(cell) if cell is not None else ""
            val_escaped = html.escape(val)
            if val == "PASS" or val == "PASSED":
                val_escaped = '<span class="badge-pass">PASS</span>'
            html_content += f"<{tag}>{val_escaped}</{tag}>\n"
        html_content += "</tr>\n"
        first = False
        
    html_content += "</table></div>\n"

html_content += """
<script>
function showSheet(id, btn) {
    document.querySelectorAll('.sheet-container').forEach(el => el.classList.remove('active'));
    document.querySelectorAll('.tab-btn').forEach(el => el.classList.remove('active'));
    document.getElementById(id).classList.add('active');
    btn.classList.add('active');
}
</script>
</body>
</html>
"""

with open(r'c:\Users\santh\AndroidStudioProjects\AeroAssistAI\backend\AeroAssist_Backend_Security_Report.html', 'w', encoding='utf-8') as f:
    f.write(html_content)

print("HTML report generated successfully!")
