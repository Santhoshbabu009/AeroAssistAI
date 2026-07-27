import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pytest
import os
import sys

# Ensure local backend imports work
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
os.environ["USE_SQLITE_TEST"] = "1"
os.environ["DISABLE_RATE_LIMIT"] = "1"

class ExcelTestCollector:
    def __init__(self, filename, title):
        self.filename = filename
        self.title = title
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)  # remove default sheet
        
    def add_suite_results(self, sheet_name, results):
        ws = self.wb.create_sheet(title=sheet_name[:31])
        
        # Styles
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        pass_font = Font(name="Segoe UI", size=10, color="006100", bold=True)
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        fail_font = Font(name="Segoe UI", size=10, color="9C0006", bold=True)
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        headers = ["Test ID", "Test Module", "Test Function", "Description", "Status", "Details"]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        for row_idx, res in enumerate(results, start=2):
            ws.append([
                res.get("id", f"TC-{row_idx-1:03d}"),
                res.get("module", ""),
                res.get("name", ""),
                res.get("doc", ""),
                res.get("status", "PASSED"),
                res.get("details", "Verified successfully.")
            ])
            
            status_cell = ws.cell(row=row_idx, column=5)
            if res.get("status") == "PASSED":
                status_cell.fill = pass_fill
                status_cell.font = pass_font
            else:
                status_cell.fill = fail_fill
                status_cell.font = fail_font
                
            for col in range(1, 7):
                ws.cell(row=row_idx, column=col).border = thin_border

        # Adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 60)

    def save(self):
        backend_dir = os.path.dirname(os.path.abspath(__file__))
        root_dir = os.path.abspath(os.path.join(backend_dir, ".."))
        
        filepath_backend = os.path.join(backend_dir, self.filename)
        filepath_root = os.path.join(root_dir, self.filename)
        
        self.wb.save(filepath_backend)
        self.wb.save(filepath_root)
        print(f"Saved Excel Report: {filepath_backend} & {filepath_root}")

def generate_all_excel_reports():
    test_files = [
        ("AeroAssist_API_Test_Report.xlsx", "backend/tests/test_api.py", "API Endpoints Suite"),
        ("AeroAssist_Auth_Test_Report.xlsx", "backend/tests/test_auth.py", "Authentication Suite"),
        ("AeroAssist_Database_Security_Report.xlsx", "backend/tests/test_database_security.py", "Database & RBAC Security Suite"),
        ("AeroAssist_Vendor_Test_Report.xlsx", "backend/tests/test_vendor.py", "Vendor & Products Suite")
    ]
    
    for filename, test_file, title in test_files:
        collector = ExcelTestCollector(filename, title)
        
        class TestPlugin:
            def __init__(self):
                self.results = []
            def pytest_runtest_logreport(self, report):
                if report.when == "call":
                    self.results.append({
                        "id": report.nodeid.split("::")[-1],
                        "module": os.path.basename(report.location[0]),
                        "name": report.nodeid.split("::")[-1],
                        "doc": report.keywords.get("__doc__", ""),
                        "status": "PASSED" if report.passed else "FAILED",
                        "details": str(report.longrepr) if report.failed else "Assertion validated successfully."
                    })
                    
        plugin = TestPlugin()
        pytest.main(["-q", test_file], plugins=[plugin])
        collector.add_suite_results("Test Results", plugin.results)
        collector.save()

if __name__ == "__main__":
    generate_all_excel_reports()

