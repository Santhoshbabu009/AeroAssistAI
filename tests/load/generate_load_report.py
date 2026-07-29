import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os

class LoadExcelReporter:
    def __init__(self, results_file: str, output_file: str = "load_report.xlsx"):
        self.results_file = results_file
        self.output_file = output_file
        self.data = []

    def parse_results(self):
        if not os.path.exists(self.results_file):
            print(f"Results file {self.results_file} not found. Generating dummy data.")
            self.data = [
                {"Scenario Name": "Load Test 10", "VUs": 10, "Duration": "3m", "Total Requests": 500, "Passed": 490, "Failed": 10, "Error Rate": 0.02, "p50 (ms)": 300, "p90 (ms)": 800, "p95 (ms)": 1500, "p99 (ms)": 2500, "Throughput (req/s)": 2.7, "Status": "PASS"},
                {"Scenario Name": "Stress Test", "VUs": 1000, "Duration": "10m", "Total Requests": 50000, "Passed": 45000, "Failed": 5000, "Error Rate": 0.1, "p50 (ms)": 2000, "p90 (ms)": 4500, "p95 (ms)": 5500, "p99 (ms)": 8000, "Throughput (req/s)": 83.3, "Status": "FAIL"}
            ]
            return

        with open(self.results_file, 'r') as f:
            for line in f:
                try:
                    entry = json.loads(line)
                    if entry.get("type") == "Point" and entry.get("metric") == "http_req_duration":
                        # Basic parsing logic here, simplified for example
                        pass
                except:
                    pass
        # Populate self.data from parsed logic

    def generate_report(self):
        self.parse_results()
        df = pd.DataFrame(self.data)
        with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Load Results', index=False)
            worksheet = writer.sheets['Load Results']

            # Conditional formatting
            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            for row in range(2, len(df) + 2):
                p95 = worksheet[f"J{row}"].value # J is p95 (ms) column
                if p95:
                    if p95 < 2000:
                        worksheet[f"J{row}"].fill = green_fill
                    elif p95 < 5000:
                        worksheet[f"J{row}"].fill = yellow_fill
                    else:
                        worksheet[f"J{row}"].fill = red_fill

if __name__ == "__main__":
    reporter = LoadExcelReporter("results.json")
    reporter.generate_report()
